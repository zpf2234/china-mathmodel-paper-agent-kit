#!/usr/bin/env python
"""Run deterministic compliance and artifact checks for a CUMCM submission."""

from __future__ import annotations

import argparse
import hashlib
import json
import math
import re
import zipfile
from pathlib import Path


SCORE_DIMENSIONS = (
    "题意与口径", "数据理解", "模型适配", "数学严谨", "求解实现", "验证强度",
    "结果价值", "证据追溯", "可复现性", "写作原创", "可视表达", "提交就绪",
)

# Year-specific official limits and disclosure policy.  Keeping this centralized
# prevents current rules from invalidating historical papers.
CONTEST_RULES = {
    2024: {"body_page_limit": 30, "size_limit_mb": 20.0, "ai_disclosure_required": False},
    2025: {"body_page_limit": 30, "size_limit_mb": 20.0, "ai_disclosure_required": True},
    2026: {"body_page_limit": 30, "size_limit_mb": 20.0, "ai_disclosure_required": True},
}


def contest_rules(year: int) -> dict[str, object]:
    """Return effective rules for *year*, using the latest known earlier rules."""
    if year in CONTEST_RULES:
        return dict(CONTEST_RULES[year])
    eligible = [known for known in CONTEST_RULES if known <= year]
    selected = max(eligible) if eligible else min(CONTEST_RULES)
    return dict(CONTEST_RULES[selected])


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def tex_formal_figures(tex: str) -> list[dict]:
    """Return captioned/labeled figure environments, independent of registries."""
    figures: list[dict] = []
    for index, match in enumerate(
        re.finditer(r"\\begin\{figure\*?\}(.*?)\\end\{figure\*?\}", tex, re.DOTALL),
        start=1,
    ):
        block = match.group(1)
        images = re.findall(r"\\includegraphics(?:\[[^\]]*\])?\{([^}]+)\}", block)
        caption = re.search(r"\\caption(?:\[[^\]]*\])?\{([^}]*)\}", block)
        label = re.search(r"\\label\{([^}]+)\}", block)
        if images and caption and label:
            figures.append(
                {
                    "index": index,
                    "label": label.group(1),
                    "caption": caption.group(1),
                    "images": images,
                    "stems": sorted({Path(value).stem for value in images}),
                }
            )
    return figures


def percentile(values: list[float], q: float) -> float | None:
    if not values:
        return None
    ordered = sorted(values)
    position = q * (len(ordered) - 1)
    lower, upper = math.floor(position), math.ceil(position)
    return ordered[lower] + (ordered[upper] - ordered[lower]) * (position - lower)


def validate_provenance(root: Path, item: dict) -> list[str]:
    generation = item.get("generation")
    if not isinstance(generation, dict):
        return ["missing generation object"]
    relative = generation.get("provenance_manifest")
    if not isinstance(relative, str) or not (root / relative).exists():
        return [f"missing provenance manifest: {relative}"]
    try:
        manifest = json.loads(read_text(root / relative))
    except Exception as exc:  # noqa: BLE001
        return [f"unreadable provenance manifest {relative}: {exc}"]
    if not isinstance(manifest, dict):
        return [f"provenance manifest root must be an object: {relative}"]
    errors: list[str] = []
    if manifest.get("status") != "VERIFIED" and manifest.get("verified") is not True:
        errors.append(f"provenance is not VERIFIED: {relative}")
    output = item.get("vector_output")
    if not isinstance(output, str) or not (root / output).exists():
        errors.append(f"provenance output missing: {output}")
        return errors
    manifest_output = manifest.get("output")
    if manifest_output and Path(str(manifest_output)).as_posix() != Path(output).as_posix():
        errors.append(f"provenance output mismatch: {relative}")
    expected_hash = manifest.get("output_sha256")
    actual_hash = sha256_file(root / output)
    if not isinstance(expected_hash, str) or expected_hash.lower() != actual_hash:
        errors.append(f"provenance output SHA-256 mismatch or missing: {relative}")
    inputs = manifest.get("inputs")
    if not isinstance(inputs, list) or not inputs:
        errors.append(f"provenance inputs missing: {relative}")
    else:
        seen_paths: set[str] = set()
        for record in inputs:
            if not isinstance(record, dict):
                errors.append(f"provenance input record invalid: {relative}")
                continue
            path_value = record.get("path")
            hash_value = record.get("sha256")
            role = record.get("role")
            if not isinstance(path_value, str) or not path_value or path_value in seen_paths:
                errors.append(f"provenance input path missing or duplicate: {relative}")
                continue
            seen_paths.add(path_value)
            input_path = root / path_value
            if not input_path.exists() or not input_path.is_file():
                errors.append(f"provenance input missing: {path_value}")
                continue
            if not isinstance(role, str) or not role:
                errors.append(f"provenance input role missing: {path_value}")
            if not isinstance(hash_value, str) or hash_value.lower() != sha256_file(input_path):
                errors.append(f"provenance input SHA-256 mismatch or missing: {path_value}")
    return errors


def validate_national_scorecard(
    root: Path, scorecard_path: Path, pdf_path: Path, pdf_sha256: str | None
) -> list[str]:
    errors: list[str] = []
    if not scorecard_path.exists():
        return ["national-first scorecard missing: 审查/评分卡.json"]
    try:
        card = json.loads(read_text(scorecard_path))
    except Exception as exc:  # noqa: BLE001
        return [f"national-first scorecard unreadable: {exc}"]
    if not isinstance(card, dict):
        return ["national-first scorecard root must be an object"]
    dimensions = card.get("dimensions")
    if not isinstance(dimensions, dict) or set(dimensions) != set(SCORE_DIMENSIONS):
        errors.append("scorecard must contain exactly the 12 required dimensions")
        dimensions = {}
    scores: dict[str, int] = {}
    for name in SCORE_DIMENSIONS:
        item = dimensions.get(name)
        score = item.get("score") if isinstance(item, dict) else None
        evidence = item.get("evidence") if isinstance(item, dict) else None
        if isinstance(item, dict) and item.get("status") == "REVIEW_REQUIRED":
            errors.append(f"scorecard dimension still requires review: {name}")
            continue
        if not isinstance(score, int) or isinstance(score, bool) or not 0 <= score <= 5:
            errors.append(f"scorecard dimension has invalid score: {name}")
            continue
        scores[name] = score
        if not isinstance(evidence, list) or not evidence:
            errors.append(f"scorecard dimension lacks evidence: {name}")
        else:
            for relative in evidence:
                if not isinstance(relative, str) or not (root / relative).exists():
                    errors.append(f"scorecard evidence missing for {name}: {relative}")
    computed_total = sum(scores.values())
    if card.get("total") != computed_total:
        errors.append(f"scorecard total mismatch: declared {card.get('total')}, computed {computed_total}")
    if computed_total < 57:
        errors.append(f"national-first score below 57/60: {computed_total}")
    if scores.get("可视表达") != 5:
        errors.append("national-first visual-expression score must be 5/5")
    if card.get("hard_gates_pass") is not True:
        errors.append("scorecard hard_gates_pass must be true")
    if card.get("p0_findings") != [] or card.get("p1_findings") != []:
        errors.append("scorecard contains unresolved P0/P1 findings")
    if card.get("verdict") != "PASS_NATIONAL_FIRST_CANDIDATE":
        errors.append("scorecard verdict must be PASS_NATIONAL_FIRST_CANDIDATE")
    if card.get("review_required"):
        errors.append("scorecard contains unresolved REVIEW_REQUIRED dimensions")
    binding = card.get("pdf_binding")
    if not isinstance(binding, dict):
        errors.append("scorecard lacks pdf_binding")
    else:
        bound_path = binding.get("path")
        bound_hash = binding.get("sha256")
        try:
            same_pdf = isinstance(bound_path, str) and (root / bound_path).resolve() == pdf_path.resolve()
        except OSError:
            same_pdf = False
        if not same_pdf:
            errors.append(f"scorecard pdf_binding.path does not identify {pdf_path.relative_to(root)}")
        if not pdf_sha256 or bound_hash != pdf_sha256:
            errors.append("scorecard PDF SHA-256 does not match the audited PDF")
    return errors


def read_text(path: Path) -> str:
    try:
        return path.read_text(encoding="utf-8")
    except UnicodeDecodeError:
        return path.read_text(encoding="gbk", errors="ignore")


def pdf_page_count(path: Path) -> int | None:
    try:
        try:
            from pypdf import PdfReader  # type: ignore
        except ImportError:
            from PyPDF2 import PdfReader  # type: ignore
        return len(PdfReader(str(path)).pages)
    except Exception:  # noqa: BLE001
        return None


def pdf_page_texts(path: Path) -> list[str] | None:
    try:
        try:
            from pypdf import PdfReader  # type: ignore
        except ImportError:
            from PyPDF2 import PdfReader  # type: ignore
        return [page.extract_text() or "" for page in PdfReader(str(path)).pages]
    except Exception:  # noqa: BLE001
        return None


def pdf_text_lengths(path: Path) -> list[int] | None:
    texts = pdf_page_texts(path)
    if texts is None:
        return None
    return [len(re.sub(r"\s+", "", text)) for text in texts]


def detect_references_start_page(page_texts: list[str]) -> int | None:
    """Find a bibliography heading backed by a reference-entry marker."""
    heading = re.compile(r"(?m)^\s*(?:第?[一二三四五六七八九十\d]+[、.．\s]*)?参考文献\s*$")
    entry = re.compile(r"(?m)^\s*(?:\[\s*\d+\s*\]|\d+[.．、])\s*\S+")
    for index, text in enumerate(page_texts):
        normalized = text.replace("\r", "")
        if not heading.search(normalized):
            continue
        evidence = normalized
        if index + 1 < len(page_texts):
            evidence += "\n" + page_texts[index + 1]
        if entry.search(evidence):
            return index + 1
    return None


def first_heading_page(page_texts: list[str], heading: str) -> int | None:
    compact = re.sub(r"\s+", "", heading)
    for page_number, text in enumerate(page_texts, start=1):
        lines = [re.sub(r"\s+", "", line) for line in text.splitlines()]
        if compact in lines:
            return page_number
    return None


def pdf_image_counts(path: Path) -> list[int] | None:
    try:
        try:
            from pypdf import PdfReader  # type: ignore
        except ImportError:
            from PyPDF2 import PdfReader  # type: ignore
        counts: list[int] = []
        for page in PdfReader(str(path)).pages:
            resources = page.get("/Resources", {})
            if hasattr(resources, "get_object"):
                resources = resources.get_object()
            xobjects = resources.get("/XObject", {}) if resources else {}
            if hasattr(xobjects, "get_object"):
                xobjects = xobjects.get_object()
            count = 0
            for raw_object in xobjects.values() if xobjects else []:
                obj = raw_object.get_object() if hasattr(raw_object, "get_object") else raw_object
                if obj.get("/Subtype") == "/Image":
                    count += 1
            counts.append(count)
        return counts
    except Exception:  # noqa: BLE001
        return None


def label_page(aux_text: str, label: str) -> int | None:
    pattern = rf"\\newlabel\{{{re.escape(label)}\}}\{{\{{[^}}]*\}}\{{(\d+)\}}"
    match = re.search(pattern, aux_text)
    return int(match.group(1)) if match else None


def strip_tex_comments(text: str) -> str:
    return re.sub(r"(?m)(?<!\\)%.*$", "", text)


def included_tex_files(main_tex: Path) -> tuple[list[Path], list[str]]:
    visited: set[Path] = set()
    ordered: list[Path] = []
    missing: list[str] = []

    def visit(path: Path) -> None:
        path = path.resolve()
        if path in visited:
            return
        if not path.exists():
            missing.append(str(path))
            return
        visited.add(path)
        ordered.append(path)
        text = strip_tex_comments(read_text(path))
        for raw_target in re.findall(r"\\(?:input|include)\{([^}]+)\}", text):
            target = (path.parent / raw_target)
            if target.suffix.lower() != ".tex":
                target = target.with_suffix(".tex")
            visit(target)

    visit(main_tex)
    return ordered, missing


def resolve_abstract_source(main_tex: Path, tex_files: list[Path]) -> tuple[Path | None, str]:
    """Resolve modular or inline abstract source without enforcing a filename."""
    preferred = main_tex.parent / "0.摘要.tex"
    candidates = ([preferred] if preferred.exists() else []) + [
        path for path in tex_files if path != preferred
    ]
    for path in candidates:
        text = read_text(path)
        if re.search(r"\\begin\{abstract\}", text):
            return path, text
    return None, ""


def audit_appendix(paper_dir: Path, tex_blob: str, appendix_files: list[Path]) -> tuple[list[str], list[str]]:
    """Audit a modular/inline appendix; absence itself is legal."""
    hard: list[str] = []
    warnings: list[str] = []
    appendix_text = "\n".join(read_text(path) for path in appendix_files)
    if not appendix_text:
        marker = re.search(r"\\appendix\b", tex_blob)
        if marker:
            appendix_text = tex_blob[marker.start():]
    if not appendix_text:
        warnings.append("no appendix detected; legal when no appendix/supporting code is required")
        return hard, warnings
    if "支撑材料" not in appendix_text:
        hard.append("appendix lacks support-material list")
    if not re.search(r"运行环境|开发环境", appendix_text):
        hard.append("appendix lacks runtime environment")
    if not re.search(r"\\label\{app:[^}]+\}", appendix_text):
        hard.append("appendix lacks stable app:* labels")
    code_inputs = re.findall(r"\\lstinputlisting(?:\[[^\]]*\])?\{([^}]+)\}", appendix_text)
    if "lstlisting" not in appendix_text and not code_inputs:
        hard.append("appendix lacks source code listing")
    missing_code = [raw for raw in code_inputs if not (paper_dir / raw).resolve().exists()]
    if missing_code:
        hard.append(f"appendix references missing source files: {', '.join(missing_code[:10])}")
    if "此处省略" in appendix_text or re.search(r"\.\.\.+", appendix_text):
        hard.append("appendix may contain omitted code")
    return hard, warnings


def audit_formula_style(tex: str) -> tuple[list[str], list[str]]:
    """Apply contextual formula style checks; retain a hard gate for true systems."""
    hard: list[str] = []
    warnings: list[str] = []
    display_start = re.compile(r"\\begin\{(equation|align|gather|multline)\}")
    for index, match in enumerate(display_start.finditer(tex), start=1):
        prefix = tex[:match.start()].rstrip()
        if not prefix.endswith(("：", ":")):
            warnings.append(f"display formula {index} is not introduced by a colon; review sentence context")
        environment = match.group(1)
        end_match = re.search(rf"\\end\{{{environment}\}}", tex[match.end():])
        if not end_match:
            continue
        block = tex[match.end():match.end() + end_match.start()]
        block = re.sub(r"\\label\{[^}]+\}", "", block)
        block = re.sub(r"\\end\{(?:aligned|split|cases)\}\s*$", "", block.strip()).rstrip()
        if re.search(r"[，。；,;]$", block) or (block.endswith(".") and not block.endswith(r"\right.")):
            warnings.append(f"display formula {index} ends with punctuation; acceptable in sentence-style mathematics")
    for index, match in enumerate(
        re.finditer(r"\\begin\{aligned\}(.*?)\\end\{aligned\}", tex, flags=re.DOTALL), start=1
    ):
        aligned_body = match.group(1)
        if re.search(r"[，。；,;]\s*\\\\", aligned_body):
            warnings.append(f"aligned formula group {index} contains row-ending punctuation")
        if aligned_body.count("&=") < 2:
            continue
        context = tex[max(0, match.start() - 80):match.start()]
        true_system = bool(re.search(r"(?:联立|方程组|约束组|方程系统)[^。；\n]{0,24}[：:]?\s*$", context))
        if true_system and r"\left\{" not in context:
            hard.append(f"true equation system {index} lacks a left brace")
    return hard, warnings


def load_registry(path: Path, collection: str) -> tuple[list[dict], str | None]:
    if not path.exists():
        return [], f"{path.name} missing"
    try:
        payload = json.loads(read_text(path))
    except Exception as exc:  # noqa: BLE001
        return [], f"{path.name} unreadable: {exc}"
    items = payload.get(collection) if isinstance(payload, dict) else None
    if not isinstance(items, list):
        return [], f"{path.name}.{collection} must be a list"
    return [item for item in items if isinstance(item, dict)], None


def normalized_artifact_stem(value: str) -> str:
    stem = Path(value).stem
    return re.sub(r"(?:_final|_preview|_actual|_page|_thumb|_gray)$", "", stem, flags=re.IGNORECASE)


def registry_match(items: list[dict], image_path: str) -> dict | None:
    target = normalized_artifact_stem(image_path)
    for item in items:
        vector = item.get("vector_output")
        if isinstance(vector, str) and normalized_artifact_stem(vector) == target:
            return item
    return None


def workbook_output_cells(path: Path) -> tuple[int | None, str | None]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError:
        return None, "openpyxl unavailable"
    try:
        workbook = load_workbook(path, read_only=True, data_only=False)
        count = 0
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    value = cell.value
                    if isinstance(value, bool):
                        continue
                    if isinstance(value, (int, float)):
                        count += 1
                    elif isinstance(value, str) and value.startswith("="):
                        count += 1
        workbook.close()
        return count, None
    except Exception as exc:  # noqa: BLE001
        return None, str(exc)


def declared_output_workbooks(root: Path, solve_dir: Path) -> list[Path]:
    contract_path = solve_dir / "任务契约.json"
    if contract_path.exists():
        try:
            contract = json.loads(read_text(contract_path))
        except Exception:  # noqa: BLE001
            contract = {}
        outputs = contract.get("官方输出", []) if isinstance(contract, dict) else []
        resolved: list[Path] = []
        for item in outputs if isinstance(outputs, list) else []:
            if not isinstance(item, dict) or not item.get("路径"):
                continue
            raw = Path(str(item["路径"]))
            candidates = [raw] if raw.is_absolute() else [
                root / raw,
                solve_dir / raw.name,
                root / "数据" / raw.name,
            ]
            path = next((candidate.resolve() for candidate in candidates if candidate.exists()), None)
            if path is not None and path.suffix.lower() == ".xlsx":
                resolved.append(path)
        return sorted(set(resolved))
    return sorted(set((root / "数据").glob("result*.xlsx")) | set(root.glob("result*.xlsx")))


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--root", default=".", help="CUMCM project root")
    parser.add_argument("--body-page-min", type=int, default=20)
    parser.add_argument("--contest-year", type=int, default=None, help="competition rule year; inferred from TeX when omitted")
    parser.add_argument(
        "--body-page-limit", "--body-page-max", dest="body_page_limit", type=int, default=None
    )
    parser.add_argument("--size-limit-mb", type=float, default=None)
    parser.add_argument(
        "--track", choices=("standard", "national-first"), default="standard",
        help="national-first enables 57/60, benchmark depth, provenance and PDF hash gates",
    )
    parser.add_argument("--no-write", action="store_true")
    args = parser.parse_args()

    root = Path(args.root).resolve()
    paper_dir = root / "论文"
    main_tex = paper_dir / "论文.tex"
    main_text = read_text(main_tex) if main_tex.exists() else ""
    year_match = re.search(r"\\yearinput\{(20\d{2})\}", main_text)
    contest_year = args.contest_year or (int(year_match.group(1)) if year_match else 2026)
    rules = contest_rules(contest_year)
    if args.body_page_limit is None:
        args.body_page_limit = int(rules["body_page_limit"])
    if args.size_limit_mb is None:
        args.size_limit_mb = float(rules["size_limit_mb"])
    if args.body_page_min < 1 or args.body_page_min > args.body_page_limit:
        parser.error("body page range must satisfy 1 <= min <= max")

    size_limit = int(args.size_limit_mb * 1024 * 1024)
    solve_dir = root / "求解"
    review_dir = root / "审查"
    review_dir.mkdir(exist_ok=True)

    hard: list[str] = []
    warnings: list[str] = []

    pdf = paper_dir / "论文.pdf"
    log = paper_dir / "论文.log"
    aux = paper_dir / "论文.aux"
    tex_files, missing_inputs = included_tex_files(main_tex)
    tex_blob = strip_tex_comments("\n".join(read_text(path) for path in tex_files))

    title_match = re.search(r"\\title\{([^{}]+)\}", read_text(main_tex))
    if not title_match:
        hard.append("paper title missing from main TeX file")
    else:
        paper_title = re.sub(r"\s+", "", title_match.group(1))
        if re.search(r"[。！？.!?]$", paper_title):
            hard.append("paper title ends with punctuation")
        if len(paper_title) < 7 or len(paper_title) > 33:
            warnings.append(
                f"paper title length {len(paper_title)} is outside the "
                "7--33 character corpus observation range; review manually"
            )
        if re.search(r"AI|人工智能|获奖|一等奖|优秀论文", paper_title, re.IGNORECASE):
            hard.append("paper title contains AI or award-oriented wording")

    pdf_sha256 = None
    if not pdf.exists() or pdf.stat().st_size == 0:
        hard.append("论文.pdf missing or empty")
    else:
        pdf_sha256 = sha256_file(pdf)
        if pdf.stat().st_size > size_limit:
            hard.append(f"论文.pdf exceeds {args.size_limit_mb:g}MB")
    total_pages = pdf_page_count(pdf) if pdf.exists() else None
    if pdf.exists() and total_pages is None:
        warnings.append("PDF page count unavailable")

    aux_text = read_text(aux) if aux.exists() else ""
    abstract_end = label_page(aux_text, "abstract:end")
    references_start = label_page(aux_text, "references:start")
    appendix_start = label_page(aux_text, "appendix:start")
    page_texts = pdf_page_texts(pdf) if pdf.exists() else None
    references_page_source = "aux-label" if references_start is not None else None
    if references_start is None and page_texts is not None:
        references_start = detect_references_start_page(page_texts)
        if references_start is not None:
            references_page_source = "pdf-heading-fallback"
            warnings.append("references:start label missing; body page boundary recovered from PDF heading")
    body_pages = references_start - 1 if references_start is not None else None
    sparse_body_pages: list[tuple[int, int]] = []
    if abstract_end is None:
        if page_texts is None:
            hard.append("abstract:end label missing; abstract page limit not machine-verified")
        else:
            warnings.append("abstract:end label missing; verify the first-page abstract boundary visually")
    elif abstract_end > 1:
        hard.append(f"abstract exceeds one page: ends on page {abstract_end}")
    if references_start is None:
        hard.append("references:start label missing and PDF reference heading fallback failed; required body page range not verified")
    elif references_page_source == "aux-label" and not re.search(r"\\clearpage\s*\\label\s*\{references:start\}", tex_blob):
        hard.append("references:start must immediately follow clearpage")
    elif body_pages < args.body_page_min:
        hard.append(
            f"body below {args.body_page_min} pages: {body_pages}"
        )
    elif body_pages > args.body_page_limit:
        hard.append(
            f"body exceeds {args.body_page_limit} pages: {body_pages}"
        )
    if pdf.exists() and body_pages is not None:
        page_text_lengths = pdf_text_lengths(pdf)
        if page_text_lengths is None:
            warnings.append("PDF text density unavailable")
        else:
            page_image_counts = pdf_image_counts(pdf)
            sparse_body_pages = [
                (page_number, page_text_lengths[page_number - 1])
                for page_number in range(1, min(body_pages, len(page_text_lengths)) + 1)
                if page_text_lengths[page_number - 1] < 180
                and (
                    page_image_counts is None
                    or page_number > len(page_image_counts)
                    or page_image_counts[page_number - 1] == 0
                )
            ]
            for page_number, char_count in sparse_body_pages:
                warnings.append(
                    f"body page {page_number} has only {char_count} extractable "
                    "characters; inspect abnormal whitespace or oversized floats"
                )

    if not log.exists():
        hard.append("论文.log missing")
    else:
        log_text = read_text(log)
        if re.search(r"(^|\n)!", log_text) or "LaTeX Error" in log_text:
            hard.append("论文.log contains LaTeX errors")
        box_warnings = re.findall(r"Overfull|Float too large", log_text)
        if box_warnings:
            warnings.append(f"layout warnings in log: {len(box_warnings)}")
        if "undefined references" in log_text.lower() or "citation" in log_text.lower() and "undefined" in log_text.lower():
            warnings.append("unresolved references or citations may remain")

    if not main_tex.exists():
        hard.append("论文.tex missing")
    if missing_inputs:
        hard.append(f"missing included tex files: {', '.join(missing_inputs[:10])}")
    if "\\tableofcontents" in tex_blob:
        hard.append("table of contents command found")
    for token in ["TODO", "待补充", "占位符", "此处补充", "可进一步补充"]:
        if token in tex_blob:
            hard.append(f"unfinished token found: {token}")

    abstract_path, abstract_text = resolve_abstract_source(main_tex, tex_files)
    if abstract_path is None:
        hard.append("abstract environment missing from included TeX sources")
    else:
        abstract_body_match = re.search(
            r"\\begin\{abstract\}(.*?)\\keywords\{",
            abstract_text,
            flags=re.DOTALL,
        )
        abstract_body = abstract_body_match.group(1) if abstract_body_match else abstract_text
        abstract_plain = re.sub(r"\\[A-Za-z]+\*?(?:\[[^\]]*\])?", "", abstract_body)
        abstract_plain = re.sub(r"[\\{}$^_&\s]", "", abstract_plain)
        abstract_chars = len(abstract_plain)
        abstract_paragraphs = len(
            [part for part in re.split(r"\n\s*\n", abstract_body) if part.strip()]
        )
        if abstract_chars < 750:
            warnings.append(f"abstract is empirically short: {abstract_chars} chars")
        elif abstract_chars > 1400:
            warnings.append(f"abstract is empirically long: {abstract_chars} chars")
        if abstract_paragraphs < 4:
            warnings.append(
                f"abstract has few natural paragraphs: {abstract_paragraphs}"
            )
        praise_words = ["效果显著", "极大应用价值", "完美解决", "非常优越", "极其精准"]
        for word in praise_words:
            if word in abstract_text:
                hard.append(f"abstract contains subjective praise: {word}")
        if not re.search(r"\d+(?:\.\d+)?", abstract_text):
            hard.append("abstract lacks quantitative numerical results")
        abstract_ai_patterns = (
            r"(?<![A-Za-z])AI(?![A-Za-z])|人工智能|提示词|prompt|"
            r"ChatGPT|Codex|OpenAI|生成过程|辅助写作|作为(?:语言|大语言)模型"
        )
        if re.search(abstract_ai_patterns, abstract_body, re.IGNORECASE):
            hard.append("abstract exposes AI prompts, tools, or generation process")
        if not re.search(r"\\keywords\{[^}]{0,80}(?:\\textbf|\\heiti|\\keyterms)", abstract_text):
            hard.append("abstract keywords are not bold")
        if not re.search(r"\\(?:textbf|keymethod|keyresult)\{", abstract_text):
            hard.append("abstract lacks selective bold emphasis for core methods/results")
        semantic_method_count = len(re.findall(r"\\keymethod\{", abstract_body))
        semantic_result_count = len(re.findall(r"\\keyresult\{", abstract_body))
        if semantic_method_count == 0 or semantic_result_count == 0:
            warnings.append(
                "abstract bold emphasis is not semantically separated into "
                "\\keymethod and \\keyresult"
            )
        bold_count = len(
            re.findall(r"\\(?:textbf|keymethod|keyresult)\{", abstract_body)
        )
        problem_count = len(
            set(re.findall(r"问题[一二三四五六七八九十]+", abstract_body))
        )
        bold_soft_limit = max(4, 2 * problem_count + 2)
        if bold_count > bold_soft_limit:
            warnings.append(
                f"abstract has {bold_count} emphasized spans; review whether bold "
                "is limited to methods and final conclusions"
            )
        keywords_match = re.search(r"\\keywords\{([^}]+)\}", abstract_text)
        if keywords_match:
            keyword_count = len(
                [item for item in re.split(r"[；;]", keywords_match.group(1)) if item.strip()]
            )
            if not 3 <= keyword_count <= 6:
                warnings.append(f"abstract keyword count outside 3-6: {keyword_count}")

    appendix_files = [path for path in tex_files if "附录" in path.stem]
    appendix_hard, appendix_warnings = audit_appendix(paper_dir, tex_blob, appendix_files)
    hard.extend(appendix_hard)
    warnings.extend(appendix_warnings)

    ai_statement_files = [path for path in tex_files if "AI" in path.stem or "人工智能" in path.stem]
    reference_files = [path for path in tex_files if "参考文献" in path.stem]
    ai_statement_text = "\n".join(read_text(path) for path in ai_statement_files)
    reference_text = "\n".join(read_text(path) for path in reference_files)
    if rules.get("ai_disclosure_required") is True and not ai_statement_files and "AI工具" not in tex_blob and "人工智能" not in tex_blob:
        hard.append("AI tool usage statement not found")
    no_ai_declaration = "未使用任何 AI 工具" in ai_statement_text or "未使用任何AI工具" in ai_statement_text
    known_ai_tools = ("Codex", "ChatGPT", "OpenAI", "DeepSeek", "Claude", "Gemini", "通义", "豆包")
    ai_used = not no_ai_declaration and any(token in ai_statement_text for token in known_ai_tools)
    if ai_used:
        body_files = [
            path
            for path in tex_files
            if path not in ai_statement_files
            and path not in reference_files
            and "附录" not in path.stem
            and path != main_tex
        ]
        body_text = strip_tex_comments("\n".join(read_text(path) for path in body_files))
        if re.search(
            r"\\aiassist\s*\{|AI\s*辅助|人工智能辅助|AI\s*生成|AI\s*提示词|OpenAI|Codex|ChatGPT",
            body_text,
            flags=re.IGNORECASE,
        ):
            hard.append("body exposes AI prompts, tools, annotations, or generation process")
        if not any(token in reference_text for token in known_ai_tools):
            hard.append("references lack AI tool entry")
        if not re.search(r"20\d{2}[-/.年]\d{1,2}[-/.月]\d{1,2}", reference_text):
            hard.append("AI tool reference lacks use date")
        ai_detail = root / "AI工具使用详情.pdf"
        if not ai_detail.exists() or ai_detail.stat().st_size == 0:
            hard.append("AI工具使用详情.pdf missing or empty")

    body_files_for_style = [
        path
        for path in tex_files
        if path not in ai_statement_files
        and path not in reference_files
        and "附录" not in path.stem
        and path != main_tex
    ]
    body_source = "\n".join(read_text(path) for path in body_files_for_style)
    if not body_source.strip():
        # Monolithic submissions keep all sections in 论文.tex.  Excluding main_tex would
        # otherwise erase the whole body and make section/figure counts falsely zero.
        body_source = tex_blob
    body_text_for_style = strip_tex_comments(body_source)
    body_text_for_style = re.split(
        r"\\end\{abstract\}", body_text_for_style, maxsplit=1
    )[-1]
    body_text_for_style = re.split(
        r"\\begin\{thebibliography\}|\\bibliography\{|\\appendix\b",
        body_text_for_style,
        maxsplit=1,
    )[0]

    body_ai_patterns = {
        "AI/prompt/tool residue": (
            r"(?<![A-Za-z])AI(?![A-Za-z])|人工智能|提示词|prompt|"
            r"ChatGPT|Codex|OpenAI|生成过程|辅助写作|作为(?:语言|大语言)模型"
        ),
        "attachment/file residue": (
            r"附件|"
            r"\.(?:xlsx|xls|csv|json)\b|(?:CSV|JSON)"
        ),
        "implementation residue": (
            r"代码|程序|脚本|运行命令|运行环境|支撑材料|"
            r"完整(?:求解)?程序见附录|见附录\s*\\ref|\\ref\{app:"
        ),
        "template closure residue": r"本问回答|本文的回答",
        "review-facing self-justification": (
            r"满足问题[一二三四五六七八九十\d]+[^。\n]{0,24}要求|"
            r"满足[^。\n]{0,20}(?:模型建立|可计算性)[^。\n]{0,12}要求"
        ),
        "argument-intent metanarrative": (
            r"前者回答[^。\n]{0,24}后者回答|"
            r"不因[^。\n]{0,30}预设结论"
        ),
        "definition-location guidance": (
            r"首次出现处定义|见下文定义|在下文定义|"
            r"未在表中列出[^。\n]{0,30}(?:定义|说明)"
        ),
    }
    for label, pattern in body_ai_patterns.items():
        if re.search(pattern, body_text_for_style, flags=re.IGNORECASE):
            hard.append(f"body contains {label}")

    for path in body_files_for_style:
        source = strip_tex_comments(read_text(path))
        if re.search(r"\\section\{问题[一二三四五六七八九十\d]+[：:]", source):
            if re.search(r"\\(?:newpage|clearpage|FloatBarrier)\b", source):
                hard.append(
                    f"problem section contains forced page break or end barrier: {path.name}"
                )

    section_matches = list(
        re.finditer(
            r"\\section\{(?P<title>[^}]+)\}"
            r"(?P<body>.*?)(?=\\section\{|$)",
            body_text_for_style,
            flags=re.DOTALL,
        )
    )
    section_titles = [match.group("title") for match in section_matches]
    if not any(title.strip() == "问题重述" for title in section_titles):
        hard.append("independent problem-restatement section missing")
    if not any(title.strip() == "问题分析" for title in section_titles):
        hard.append("independent problem-analysis section missing")
    if any("问题重述" in title and "问题分析" in title for title in section_titles):
        hard.append("problem restatement and analysis must be separate sections")
    if any("计算口径" in title for title in section_titles):
        hard.append("section title contains explanatory wording: 计算口径")
    symbol_index = next(
        (index for index, title in enumerate(section_titles) if "符号" in title),
        None,
    )
    assumption_index = next(
        (index for index, title in enumerate(section_titles) if "假设" in title),
        None,
    )
    if (
        symbol_index is not None
        and assumption_index is not None
        and symbol_index < assumption_index
    ):
        warnings.append(
            "the symbol section appears before model assumptions; default order "
            "is assumptions before symbol definitions unless explicitly justified"
        )

    subsection_titles = re.findall(
        r"\\subsection\{([^}]+)\}",
        body_text_for_style,
    )
    if len(subsection_titles) > 24:
        warnings.append(
            f"body contains {len(subsection_titles)} subsections; run a "
            "minimum-sufficient-heading merge review"
        )
    for section_match in section_matches:
        subsection_count = len(
            re.findall(r"\\subsection\{", section_match.group("body"))
        )
        if subsection_count > 5:
            warnings.append(
                f"section '{section_match.group('title')}' contains "
                f"{subsection_count} subsections; justify or merge headings"
            )
    for level, title in re.findall(
        r"\\(section|subsection|subsubsection)\{([^}]+)\}",
        body_text_for_style,
    ):
        compact_title = re.sub(r"\\[A-Za-z]+\*?(?:\[[^\]]*\])?", "", title)
        compact_title = re.sub(r"[{}$^_&\s：:、，,。；;]+", "", compact_title)
        if len(compact_title) > 20:
            warnings.append(
                f"long {level} title ({len(compact_title)} chars): {title}"
            )

    symbol_sections = [
        match
        for match in section_matches
        if "符号" in match.group("title")
        and not re.search(r"问题|假设|模型", match.group("title"))
    ]
    if not symbol_sections:
        hard.append("independent symbol section missing")
    else:
        symbol_body = symbol_sections[0].group("body")
        symbol_lines = [
            re.sub(r"\\[A-Za-z]+\*?(?:\[[^\]]*\])?", "", line)
            .replace("{", "")
            .replace("}", "")
            for line in symbol_body.splitlines()
            if line.count("&") >= 2
        ]
        has_three_column_header = any(
            "符号" in line
            and re.search(r"含义|意义", line)
            and re.search(r"单位|量纲", line)
            for line in symbol_lines
        )
        if not has_three_column_header:
            hard.append(
                "symbol section lacks a symbol-meaning-unit three-column header"
            )

    validation_sections = [
        match
        for match in section_matches
        if re.search(r"验证|检验|稳健|灵敏度|误差", match.group("title"))
        and "问题" not in match.group("title")
    ]
    if not validation_sections:
        hard.append("independent comprehensive validation section missing")
    else:
        validation_text = "\n".join(
            match.group("body") for match in validation_sections
        )
        validation_evidence = {
            "sensitivity/scenario": r"灵敏度|参数微扰|参数扰动|情景分析|随机噪声",
            "convergence": r"收敛|步长|网格加密|样本量|迭代次数",
            "feasibility/boundary": r"约束余量|可行性|守恒|边界检查|极端情形",
            "independent check": r"独立复算|独立实现|交叉验证|留出验证|基线比较",
            "error/scope": r"误差来源|误差传播|不确定性|适用范围|适用边界",
        }
        evidence_found = [
            label
            for label, pattern in validation_evidence.items()
            if re.search(pattern, validation_text)
        ]
        if len(evidence_found) < 2:
            hard.append(
                "comprehensive validation section contains fewer than two "
                "cross-problem evidence types"
            )

    generic_titles = {
        "模型的建立",
        "模型建立",
        "算法的实施",
        "算法实施",
        "计算结果",
        "结果显示",
        "进一步讨论",
        "一些分析",
        "相关说明",
    }
    for level, title in re.findall(
        r"\\(section|subsection)\{([^}]+)\}",
        body_text_for_style,
    ):
        compact_title = re.sub(r"[\s：:、，,。；;]+", "", title)
        if compact_title in generic_titles:
            hard.append(
                f"generic {level} title lacks problem information: {title}"
            )

    if re.search(
        r"\\(?:subsubsubsection|subparagraph)\{",
        body_text_for_style,
    ):
        hard.append("numbered heading depth exceeds three levels")

    if re.search(
        r"如(?:上|下)(?:图|表)|(?:上|下)(?:图|表)所示",
        body_text_for_style,
    ):
        hard.append("figure/table reference uses a relative position instead of a number")

    figure_registry, figure_registry_error = load_registry(
        review_dir / "figure-registry.json", "figures"
    )
    diagram_registry, diagram_registry_error = load_registry(
        review_dir / "diagram-registry.json", "diagrams"
    )
    if figure_registry_error:
        warnings.append(figure_registry_error)
    if diagram_registry_error:
        warnings.append(diagram_registry_error)

    diagram_include_pattern = re.compile(
        r"\\includegraphics(?P<options>\[[^\]]*\])?\{(?P<path>[^}]*(?:示意|流程|关系|机理|框图)[^}]*)\}"
    )
    detected_diagrams = list(diagram_include_pattern.finditer(body_text_for_style))
    # Standard submissions may predate registries; do not turn that optional metadata into
    # a hard failure.  Once a registry exists, validate it.  National-first always requires it.
    audit_diagrams = (
        detected_diagrams
        if args.track == "national-first" or diagram_registry_error is None
        else []
    )
    if detected_diagrams and not audit_diagrams:
        warnings.append(
            f"detected {len(detected_diagrams)} diagram-like figures without a usable registry; "
            "strict layout/provenance checks skipped in standard track"
        )
    for diagram_match in audit_diagrams:
        options = diagram_match.group("options") or ""
        diagram_path = Path(diagram_match.group("path"))
        if re.search(r"\b(?:trim|clip)\b", options):
            hard.append(
                f"diagram uses LaTeX trim/clip instead of a fixed source canvas: {diagram_path}"
            )
        registry_item = registry_match(diagram_registry, str(diagram_path))
        if registry_item is None:
            hard.append(f"diagram missing from diagram-registry.json: {diagram_path.stem}")
            continue
        layout_rel = registry_item.get("layout")
        layout_report = root / layout_rel if isinstance(layout_rel, str) else None
        if layout_report is None or not layout_report.exists():
            hard.append(f"diagram layout report missing: {layout_rel}")
            continue
        try:
            layout_payload = json.loads(read_text(layout_report))
        except Exception:  # noqa: BLE001
            hard.append(f"diagram layout report is unreadable: {layout_report.name}")
            continue
        if not layout_payload.get("pass", False):
            hard.append(f"diagram layout audit failed: {layout_report.name}")
        layout_checks = layout_payload.get("checks", {})
        required_layout_checks = {
            "objects_inside_canvas",
            "node_text_embedded",
            "text_inside_parent",
            "node_lines_max_2",
            "uniform_node_text_style",
            "decision_lines_max_2",
            "annotation_anchors",
            "leader_endpoint_on_target",
            "arrow_endpoints_on_node_boundary",
            "arrowheads_outside_node_interior",
            "connection_inventory_complete",
            "non_target_crossings_zero",
            "branch_label_clearance",
            "final_render_connections_checked",
            "line_direction_semantics_complete",
            "arrow_direction_matches_meaning",
            "line_style_matches_meaning",
            "text_line_clearance",
            "label_object_clearance",
            "final_font_size_readable",
            "composition_balance",
            "latex_trim_required",
        }
        if not required_layout_checks.issubset(layout_checks):
            hard.append(f"diagram layout report is incomplete: {layout_report.name}")
        elif layout_checks.get("latex_trim_required") is not False:
            hard.append(f"diagram source still requires trimming: {layout_report.name}")
        connection_audit = layout_payload.get("connection_audit", {})
        connection_items = connection_audit.get("items", [])
        connection_total = connection_audit.get("total")
        connection_checked = connection_audit.get("checked")
        connection_failed = connection_audit.get("failed")
        if (
            not isinstance(connection_items, list)
            or not isinstance(connection_total, int)
            or connection_total <= 0
            or connection_checked != connection_total
            or len(connection_items) != connection_total
            or connection_failed != 0
        ):
            hard.append(f"diagram connection audit is incomplete: {layout_report.name}")
        else:
            item_keys = {
                "id",
                "kind",
                "source",
                "target",
                "source_anchor",
                "target_anchor",
                "directionality",
                "direction",
                "meaning",
                "arrowhead_required",
                "arrowhead_present",
                "semantics_match",
                "source_endpoint_error_pt",
                "target_endpoint_error_pt",
                "non_target_crossings",
                "min_clearance_pt",
                "label_overlap",
                "node_interior_overlap",
                "final_render_pass",
            }
            for item in connection_items:
                if not isinstance(item, dict) or not item_keys.issubset(item):
                    hard.append(f"diagram connection item is incomplete: {layout_report.name}")
                    break
                if (
                    item["source_endpoint_error_pt"] > 0.5
                    or item["target_endpoint_error_pt"] > 0.5
                    or item["non_target_crossings"] != 0
                    or item["min_clearance_pt"] < 2.5
                    or item["label_overlap"] is not False
                    or item["node_interior_overlap"] is not False
                    or item["semantics_match"] is not True
                    or item["directionality"] not in {"directed", "undirected", "bidirectional"}
                    or item["arrowhead_required"] is not item["arrowhead_present"]
                    or item["final_render_pass"] is not True
                ):
                    hard.append(f"diagram connection item failed: {layout_report.name}")
                    break
        line_semantics = layout_payload.get("line_semantics_audit", {})
        line_items = line_semantics.get("items", [])
        line_total = line_semantics.get("total")
        line_checked = line_semantics.get("checked")
        line_failed = line_semantics.get("failed")
        line_item_keys = {
            "id",
            "kind",
            "meaning",
            "directionality",
            "expected_direction",
            "arrowhead_required",
            "arrowhead_present",
            "semantics_match",
            "final_render_pass",
        }
        if (
            not isinstance(line_items, list)
            or not isinstance(line_total, int)
            or line_total <= 0
            or line_checked != line_total
            or len(line_items) != line_total
            or line_failed != 0
        ):
            hard.append(f"diagram line semantics audit is incomplete: {layout_report.name}")
        else:
            for item in line_items:
                if not isinstance(item, dict) or not line_item_keys.issubset(item):
                    hard.append(f"diagram line semantics item is incomplete: {layout_report.name}")
                    break
                if (
                    item["directionality"] not in {"directed", "undirected", "bidirectional"}
                    or item["arrowhead_required"] is not item["arrowhead_present"]
                    or item["semantics_match"] is not True
                    or item["final_render_pass"] is not True
                ):
                    hard.append(f"diagram line semantics item failed: {layout_report.name}")
                    break
        overall_audit = layout_payload.get("overall_audit", {})
        required_overall_checks = {
            "pass",
            "main_axis_clear",
            "visual_hierarchy_clear",
            "spacing_consistent",
            "whitespace_balanced",
            "thumbnail_readable",
            "full_page_render_checked",
        }
        if (
            not isinstance(overall_audit, dict)
            or not required_overall_checks.issubset(overall_audit)
            or any(overall_audit.get(key) is not True for key in required_overall_checks)
        ):
            hard.append(f"diagram overall audit failed: {layout_report.name}")

    unnumbered_display = re.search(
        r"\\\[|\\begin\{(?:equation|align|gather|multline)\*\}",
        body_text_for_style,
    )
    if unnumbered_display:
        hard.append("unnumbered display formula found in body")

    formula_hard, formula_warnings = audit_formula_style(body_text_for_style)
    hard.extend(formula_hard)
    warnings.extend(formula_warnings)

    if re.search(r"\\paragraph\{[^}]*[。.]\}", body_text_for_style):
        hard.append("paragraph lead-in ends with a period instead of a colon")
    if re.search(
        r"\\item\s+\\textbf\{[^}]*[。.]\}",
        body_text_for_style,
    ):
        hard.append("bold list lead-in ends with a period instead of a colon")

    meta_patterns = {
        "便于": r"便于",
        "有助于": r"有助于",
        "彰显": r"彰显",
        "本文报告": r"本文报告",
        "不声明": r"不声明",
        "不构成证明": r"不构成[^。\n]{0,12}证明",
        "不能据此宣称": r"不能据此宣称",
    }
    for label, pattern in meta_patterns.items():
        if re.search(pattern, body_text_for_style):
            hard.append(f"body contains evaluative/meta prose: {label}")

    flow_sections = list(
        re.finditer(
            r"\\(?:section|subsection|subsubsection)"
            r"\{(?P<title>[^}]*(?:流程|五问关系与共享模型)[^}]*)\}"
            r"(?P<body>.*?)(?=\\(?:section|subsection|subsubsection)\{|$)",
            body_text_for_style,
            flags=re.DOTALL,
        )
    )
    pipeline_vocabulary = (
        "任务契约",
        "结果文件",
        "结构化指标",
        "Excel",
        "支撑材料",
        "模板检查",
        "结果交付",
    )
    for flow_match in flow_sections:
        flow_title = flow_match.group("title")
        flow_section = flow_match.group("body")
        if "\\includegraphics" not in flow_section:
            hard.append(f"flow section lacks a formal flowchart: {flow_title}")
        pipeline_terms = {
            term
            for term in pipeline_vocabulary
            if term in flow_section
        }
        if len(pipeline_terms) >= 2:
            hard.append(
                f"flow section is an artifact-production pipeline ({flow_title}): "
                + ", ".join(sorted(pipeline_terms))
            )

        flow_images = re.findall(
            r"\\includegraphics(?:\[[^\]]*\])?\{([^}]+)\}", flow_section
        )
        figure_inventory = root / "求解" / "图表清单.md"
        if flow_images and figure_inventory.exists():
            inventory_text = figure_inventory.read_text(encoding="utf-8", errors="ignore")
            for image_path in flow_images:
                image_stem = Path(image_path).stem
                if image_stem not in inventory_text:
                    hard.append(
                        f"flowchart missing from 求解/图表清单.md: {image_stem}"
                    )
                source_candidates = [
                    path
                    for path in (root / "求解").rglob(f"*{image_stem}*")
                    if path.suffix.lower() in {".py", ".m", ".tex", ".tikz", ".vsdx", ".pptx", ".drawio"}
                ]
                if not source_candidates:
                    hard.append(f"flowchart lacks editable source: {image_stem}")
                    continue
                source_pipeline_terms = set()
                for source_path in source_candidates:
                    if source_path.suffix.lower() not in {".py", ".tex", ".drawio"}:
                        continue
                    source_text = source_path.read_text(
                        encoding="utf-8", errors="ignore"
                    )
                    source_pipeline_terms.update(
                        term for term in pipeline_terms if term in source_text
                    )
                    source_pipeline_terms.update(
                        term for term in pipeline_vocabulary if term in source_text
                    )
                if len(source_pipeline_terms) >= 2:
                    hard.append(
                        "flowchart source is an artifact-production pipeline: "
                        + ", ".join(sorted(source_pipeline_terms))
                    )

    identity_patterns = {
        "phone number": r"(?<!\d)1[3-9]\d{9}(?!\d)",
        "email": r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}",
        "identity label": r"参赛队员|指导教师|联系电话|联系邮箱|所属学校|参赛学校",
    }
    for label, pattern in identity_patterns.items():
        if re.search(pattern, tex_blob):
            hard.append(f"possible identity leak: {label}")

    graphic_dirs = [paper_dir]
    for block in re.findall(r"\\graphicspath\{((?:\{[^{}]*\})+)\}", tex_blob):
        graphic_dirs.extend((paper_dir / raw).resolve() for raw in re.findall(r"\{([^{}]+)\}", block))
    missing_images = []
    for match in re.findall(r"\\includegraphics(?:\[[^\]]*\])?\{([^}]+)\}", tex_blob):
        if not any((directory / match).resolve().exists() for directory in graphic_dirs):
            missing_images.append(match)
    if missing_images:
        hard.append(f"missing referenced images: {', '.join(missing_images[:10])}")

    if not (solve_dir / "证据矩阵.csv").exists():
        hard.append("求解/证据矩阵.csv missing")
    problem_dirs = sorted(path for path in solve_dir.glob("问题*") if path.is_dir())
    metrics = []
    for problem_dir in problem_dirs:
        metrics_path = problem_dir / "结果" / "metrics.json"
        if metrics_path.exists():
            metrics.append(metrics_path)
        else:
            hard.append(f"{problem_dir.name} lacks 结果/metrics.json")
    if not problem_dirs:
        hard.append("no 求解/问题X directories")

    evidence_json = solve_dir / "证据审计.json"
    if evidence_json.exists():
        try:
            evidence_result = json.loads(read_text(evidence_json))
            if not evidence_result.get("pass"):
                hard.append("solve-stage evidence audit failed")
        except Exception as exc:  # noqa: BLE001
            hard.append(f"evidence audit report unreadable: {exc}")
    else:
        warnings.append("solve-stage evidence audit report missing")

    tex_figures = tex_formal_figures(body_text_for_style)
    registry_stems = {
        normalized_artifact_stem(str(item.get("vector_output")))
        for item in figure_registry + diagram_registry
        if isinstance(item.get("vector_output"), str)
    }
    registered_tex_figures = [
        item
        for item in tex_figures
        if any(normalized_artifact_stem(stem) in registry_stems for stem in item["stems"])
    ]
    # Registry absence must never silently turn a real TeX figure count into zero.  In the
    # standard track, count captioned+labeled TeX figures.  The national-first track counts
    # only registry-backed figures and separately blocks incomplete coverage/provenance.
    formal_figures = (
        len(registered_tex_figures) if args.track == "national-first" else len(tex_figures)
    )
    unregistered_tex_figures = [
        item for item in tex_figures if item not in registered_tex_figures
    ]
    if args.track == "national-first":
        if figure_registry_error and diagram_registry_error:
            hard.append("national-first visual registries are both missing or invalid")
        if unregistered_tex_figures:
            hard.append(
                "national-first formal figures lack registry coverage: "
                + ", ".join(item["label"] for item in unregistered_tex_figures[:20])
            )
        for registry_name, items in (
            ("figure-registry.json", figure_registry),
            ("diagram-registry.json", diagram_registry),
        ):
            for item in items:
                for error in validate_provenance(root, item):
                    hard.append(
                        f"national-first {registry_name} provenance failed for {item.get('id')}: {error}"
                    )
        figure_style_report = review_dir / "figure-style-audit.json"
        if figure_registry and not figure_style_report.exists():
            hard.append("national-first figure style/signature audit report missing")
        elif figure_registry:
            try:
                style_payload = json.loads(read_text(figure_style_report))
            except Exception as exc:  # noqa: BLE001
                hard.append(f"national-first figure style/signature audit unreadable: {exc}")
            else:
                signature = style_payload.get("signature", {}) if isinstance(style_payload, dict) else {}
                if style_payload.get("pass") is not True or style_payload.get("track") != "national-first":
                    hard.append("national-first figure style/signature audit did not pass in national-first track")
                if signature.get("status") not in {"PASS", "EXEMPTED_NO_FORCED_COLLAGE"}:
                    hard.append("national-first signature figure policy is unresolved")
    figure_density = (
        formal_figures / body_pages
        if isinstance(body_pages, int) and body_pages > 0
        else None
    )
    density_floor = 0.50
    benchmark_density_path = review_dir / "同题优秀论文基准.json"
    benchmark_q1 = None
    benchmark_median = None
    benchmark_body_q1 = None
    benchmark_count = 0
    if benchmark_density_path.exists():
        try:
            benchmark_rows = json.loads(read_text(benchmark_density_path))
            rows = [row for row in benchmark_rows if isinstance(row, dict)]
            densities = [float(row["figure_density"]) for row in rows if row.get("figure_density") is not None]
            body_benchmarks = [float(row["body_pages"]) for row in rows if row.get("body_pages") is not None]
            benchmark_count = len(rows)
            benchmark_q1 = percentile(densities, 0.25)
            benchmark_median = percentile(densities, 0.50)
            benchmark_body_q1 = percentile(body_benchmarks, 0.25)
            if benchmark_q1 is not None:
                density_floor = max(0.50, 0.90 * benchmark_q1)
        except Exception as exc:  # noqa: BLE001
            message = f"same-problem benchmark unreadable: {exc}"
            (hard if args.track == "national-first" else warnings).append(message)
    elif args.track == "national-first":
        hard.append("national-first same-problem benchmark missing")
    if args.track == "national-first" and benchmark_count < 3:
        hard.append(f"national-first benchmark sample below 3 papers: {benchmark_count}")
    minimum_body_pages = (
        math.ceil(0.90 * benchmark_body_q1) if benchmark_body_q1 is not None else None
    )
    if (
        args.track == "national-first"
        and minimum_body_pages is not None
        and isinstance(body_pages, int)
        and body_pages < minimum_body_pages
    ):
        hard.append(
            f"body depth below national-first benchmark floor: {body_pages} < {minimum_body_pages}"
        )
    required_figures = (
        math.ceil(body_pages * density_floor)
        if isinstance(body_pages, int) and body_pages > 0
        else None
    )
    density_status = "UNAVAILABLE"
    if figure_density is not None:
        density_status = "PASS" if figure_density >= density_floor else "BLOCK_FIGURE_DENSITY"
        if density_status != "PASS" and args.track == "national-first":
            hard.append(
                f"figure density below national-first floor: {figure_density:.3f} < {density_floor:.3f}"
            )
        elif density_status != "PASS":
            warnings.append(
                f"figure density below national-first reference (standard track only): "
                f"{figure_density:.3f} < {density_floor:.3f}"
            )
    density_report = {
        "track": args.track,
        "body_pages": body_pages,
        "minimum_body_pages": minimum_body_pages,
        "tex_formal_figures": len(tex_figures),
        "registry_backed_formal_figures": len(registered_tex_figures),
        "formal_figures": formal_figures,
        "unregistered_labels": [item["label"] for item in unregistered_tex_figures],
        "density": figure_density,
        "benchmark_count": benchmark_count,
        "benchmark_body_pages_q1": benchmark_body_q1,
        "benchmark_q1": benchmark_q1,
        "benchmark_median": benchmark_median,
        "floor": density_floor,
        "required_figures": required_figures,
        "status": density_status,
    }
    if not args.no_write:
        (review_dir / "图密度审查.json").write_text(
            json.dumps(density_report, ensure_ascii=False, indent=2), encoding="utf-8"
        )

    deliverables = declared_output_workbooks(root, solve_dir)
    for workbook in deliverables:
        output_cells, workbook_error = workbook_output_cells(workbook)
        if workbook_error:
            warnings.append(f"cannot inspect output workbook {workbook.name}: {workbook_error}")
        elif output_cells == 0:
            hard.append(f"required output workbook has no numeric/formula cells: {workbook.name}")

    benchmark_json = review_dir / "优秀论文对标.json"
    similarity_status = None
    if benchmark_json.exists():
        try:
            benchmark = json.loads(read_text(benchmark_json))
            similarity_status = benchmark.get("similarity", {}).get("status")
            if similarity_status == "FAIL_HIGH_SIMILARITY":
                hard.append("excellent-paper similarity audit failed")
            elif similarity_status == "WARN_REVIEW":
                warnings.append("excellent-paper similarity requires manual review")
        except Exception as exc:  # noqa: BLE001
            warnings.append(f"benchmark report unreadable: {exc}")
    else:
        warnings.append("excellent-paper benchmark/similarity report missing")

    content_review_json = review_dir / "内容审查.json"
    content_review_pass = False
    if content_review_json.exists():
        try:
            content_review = json.loads(read_text(content_review_json))
            sections_reviewed = content_review.get("sections_reviewed")
            total_sections_reviewed = content_review.get("total_sections")
            content_review_pass = (
                content_review.get("full_text_read") is True
                and isinstance(sections_reviewed, int)
                and isinstance(total_sections_reviewed, int)
                and sections_reviewed == total_sections_reviewed
                and total_sections_reviewed > 0
                and content_review.get("p0_findings") == []
                and content_review.get("p1_findings") == []
            )
            if not content_review.get("full_text_read"):
                (hard if args.track == "national-first" else warnings).append(
                    "content review does not confirm full-text reading"
                )
            if (
                not isinstance(sections_reviewed, int)
                or not isinstance(total_sections_reviewed, int)
                or sections_reviewed != total_sections_reviewed
                or total_sections_reviewed <= 0
            ):
                (hard if args.track == "national-first" else warnings).append(
                    "content review does not cover every section with numeric counts"
                )
            if content_review.get("p0_findings") or content_review.get("p1_findings"):
                (hard if args.track == "national-first" else warnings).append(
                    "content review contains unresolved P0/P1 findings"
                )
        except Exception as exc:  # noqa: BLE001
            (hard if args.track == "national-first" else warnings).append(
                f"content review report unreadable: {exc}"
            )
    else:
        (hard if args.track == "national-first" else warnings).append(
            "full-text content review report missing"
        )

    if args.track == "national-first":
        hard.extend(validate_national_scorecard(root, review_dir / "评分卡.json", pdf, pdf_sha256))

    archives = list(root.glob("*.zip")) + list(root.glob("*.rar"))
    if not archives:
        hard.append("support archive missing")
    required_template_entries = {
        f"数据/{path.name}" for path in (root / "数据").glob("result*.xlsx")
    }
    for archive in archives:
        if archive.stat().st_size > size_limit:
            hard.append(
                f"support archive exceeds {args.size_limit_mb:g}MB: {archive.name}"
            )
        if archive.suffix.lower() == ".zip":
            try:
                with zipfile.ZipFile(archive) as handle:
                    names = handle.namelist()
                normalized_names = {name.replace("\\", "/").lstrip("./") for name in names}
                if any("承诺书" in name or "编号专用" in name for name in normalized_names):
                    hard.append(f"support archive contains forbidden front matter: {archive.name}")
                missing_templates = sorted(required_template_entries - normalized_names)
                if missing_templates:
                    hard.append(
                        f"support archive lacks required template inputs: {', '.join(missing_templates)}"
                    )
                if ai_used and not any(
                    Path(name).name == "AI工具使用详情.pdf" for name in normalized_names
                ):
                    hard.append(f"support archive lacks AI工具使用详情.pdf: {archive.name}")
            except zipfile.BadZipFile:
                hard.append(f"bad zip file: {archive.name}")

    result = {
        "pass": not hard,
        "hard_errors": hard,
        "warnings": warnings,
        "counts": {
            "total_pdf_pages": total_pages,
            "abstract_end_page": abstract_end,
            "references_start_page": references_start,
            "appendix_start_page": appendix_start,
            "track": args.track,
            "contest_year": contest_year,
            "pdf_sha256": pdf_sha256,
            "body_pages": body_pages,
            "body_page_definition": "reference heading start page - 1",
            "references_page_source": references_page_source,
            "required_body_page_min": args.body_page_min,
            "required_body_page_max": args.body_page_limit,
            "appendix_and_after_page_limit": None,
            "ai_used": ai_used,
            "support_archives": len(archives),
            "metrics_json": len(metrics),
            "problem_directories": len(problem_dirs),
            "included_tex_files": len(tex_files),
            "output_workbooks": len(deliverables),
            "body_subsections": len(subsection_titles),
            "sparse_body_pages": len(sparse_body_pages),
            "formal_figures": formal_figures,
            "figure_density": figure_density,
            "figure_density_floor": density_floor,
            "required_figures": required_figures,
        },
        "figure_density": density_report,
        "similarity_status": similarity_status,
    }
    if args.no_write:
        print(json.dumps(result, ensure_ascii=False))
    else:
        (review_dir / "自动审查.json").write_text(
            json.dumps(result, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

        lines = ["# 自动审查报告", "", f"结论：{'PASS' if result['pass'] else 'FAIL'}", ""]
        lines.append("## 硬错误")
        lines.extend([f"- {item}" for item in hard] or ["- 无"])
        lines.extend(["", "## 警告"])
        lines.extend([f"- {item}" for item in warnings] or ["- 无"])
        lines.extend(["", "## 计数"])
        for key, value in result["counts"].items():
            lines.append(f"- {key}: {value}")
        (review_dir / "自动审查.md").write_text("\n".join(lines), encoding="utf-8")

    print("PASS" if result["pass"] else "FAIL")
    if not args.no_write:
        print(f"Wrote {review_dir / '自动审查.md'}")
    return 0 if result["pass"] else 1


if __name__ == "__main__":
    raise SystemExit(main())
