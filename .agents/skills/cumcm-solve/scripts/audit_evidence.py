#!/usr/bin/env python
"""Audit the solve-stage evidence package for excellent-paper readiness."""

from __future__ import annotations

import argparse
import ast
import csv
import json
from pathlib import Path


IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".pdf", ".svg"}
TABLE_EXTS = {".csv", ".xlsx", ".xls", ".json"}
VALIDATION_TOKENS = ("验证", "检验", "误差", "灵敏度", "稳健", "收敛", "残差", "可行性")
CONTRACT_FIELDS = {
    "竞赛年份",
    "赛道",
    "赛题标题",
    "题面来源",
    "附件",
    "官方输出",
    "问题",
    "参考答案在求解期可见",
}
THIRD_PARTY_IMPORTS = {
    "cvxpy": {"cvxpy"},
    "matplotlib": {"matplotlib"},
    "networkx": {"networkx"},
    "numpy": {"numpy"},
    "openpyxl": {"openpyxl"},
    "ortools": {"ortools"},
    "pandas": {"pandas"},
    "pulp": {"pulp"},
    "scipy": {"scipy"},
    "seaborn": {"seaborn"},
    "shapely": {"shapely"},
    "sklearn": {"sklearn", "scikit-learn"},
    "statsmodels": {"statsmodels"},
}


def load_json(path: Path) -> tuple[dict | None, str]:
    if not path.exists():
        return None, "missing"
    if path.stat().st_size == 0:
        return None, "empty"
    try:
        value = json.loads(path.read_text(encoding="utf-8"))
    except Exception as exc:  # noqa: BLE001
        return None, f"invalid json: {exc}"
    if not isinstance(value, dict):
        return None, "root is not an object"
    return value, "ok"


def has_numeric_result(value: object) -> bool:
    if isinstance(value, bool):
        return False
    if isinstance(value, (int, float)):
        return True
    if isinstance(value, list):
        return any(has_numeric_result(item) for item in value)
    if isinstance(value, dict):
        return any(has_numeric_result(item) for item in value.values())
    return False


def validation_present(metrics: dict) -> bool:
    for key, value in metrics.items():
        if any(token in str(key) for token in VALIDATION_TOKENS) and value not in (None, "", [], {}):
            return True
    return False


def audit_model_selection(selection: object) -> list[str]:
    issues: list[str] = []
    if not isinstance(selection, dict):
        return ["metrics.json lacks 模型选择 object"]
    mode = selection.get("模式")
    if mode not in {"候选比较", "唯一机理"}:
        issues.append("模型选择.模式 must be 候选比较 or 唯一机理")
    basis = selection.get("选择依据")
    if not isinstance(basis, list) or not basis or any(not str(item).strip() for item in basis):
        issues.append("模型选择.选择依据 requires a nonempty list")
    presentation = selection.get("论文呈现")
    if presentation not in {"表格", "段落", "简述", "不写"}:
        issues.append("模型选择.论文呈现 invalid")
    if presentation == "不写" and not str(selection.get("不写理由", "")).strip():
        issues.append("模型选择选择不写时必须给出不写理由")

    if mode == "候选比较":
        candidates = selection.get("候选")
        if not isinstance(candidates, list) or len(candidates) < 2:
            issues.append("候选比较 requires at least two candidates")
            return issues
        names: list[str] = []
        decisions: list[str] = []
        required = {"名称", "角色", "判定", "证据"}
        for index, candidate in enumerate(candidates, start=1):
            if not isinstance(candidate, dict):
                issues.append(f"模型选择候选 {index} is not an object")
                continue
            missing = sorted(required - set(candidate))
            if missing:
                issues.append(f"模型选择候选 {index} 缺字段: {', '.join(missing)}")
            name = str(candidate.get("名称", "")).strip()
            if not name:
                issues.append(f"模型选择候选 {index} 名称为空")
            names.append(name)
            decision = str(candidate.get("判定", "")).strip()
            decisions.append(decision)
            evidence = candidate.get("证据")
            if not isinstance(evidence, list) or not evidence:
                issues.append(f"模型选择候选 {index} 缺非空证据")
        if len(set(names)) != len(names):
            issues.append("模型选择候选名称重复")
        selected = str(selection.get("主模型", "")).strip()
        if not selected or selected not in names:
            issues.append("模型选择.主模型 must match a candidate")
        if decisions.count("选用") != 1:
            issues.append("候选比较 must contain exactly one 选用 decision")
        if not any(item in {"保留为验证", "淘汰"} for item in decisions):
            issues.append("候选比较 lacks an independent challenger or rejected route")
    elif mode == "唯一机理":
        if not str(selection.get("唯一性理由", "")).strip():
            issues.append("唯一机理 requires 唯一性理由")
        if not str(selection.get("独立挑战", "")).strip():
            issues.append("唯一机理 requires 独立挑战")
    return issues


def audit_answer_sufficiency(value: object) -> list[str]:
    issues: list[str] = []
    if not isinstance(value, dict):
        return ["metrics.json lacks 答案充分性 object"]
    required = (
        "题面闭环",
        "物理或约束可行",
        "独立验证",
        "数值分辨率",
        "替代路线挑战",
        "不确定性或最优性",
    )
    for field in required:
        item = value.get(field)
        if not isinstance(item, dict):
            issues.append(f"答案充分性.{field} must be an object")
            continue
        if item.get("通过") is not True:
            issues.append(f"答案充分性.{field} 未通过")
        if not str(item.get("证据", "")).strip():
            issues.append(f"答案充分性.{field} 缺证据")
    precision = value.get("物理闭合与精度预算")
    if precision is not None:
        if not isinstance(precision, dict):
            issues.append("答案充分性.物理闭合与精度预算 must be an object")
        else:
            required_precision = {
                "通过", "控制方程来源", "守恒或约束指标", "阈值", "结果",
                "单位", "合成不确定度", "报告有效位", "证据",
            }
            missing = sorted(required_precision - set(precision))
            if missing:
                issues.append("答案充分性.物理闭合与精度预算 缺字段: " + ", ".join(missing))
            if precision.get("通过") is not True:
                issues.append("答案充分性.物理闭合与精度预算 未通过")
            for field in ("阈值", "结果"):
                number = precision.get(field)
                if not isinstance(number, (int, float)) or isinstance(number, bool):
                    issues.append(f"答案充分性.物理闭合与精度预算.{field} 必须为数值")
            threshold = precision.get("阈值")
            result = precision.get("结果")
            if (
                isinstance(threshold, (int, float)) and not isinstance(threshold, bool)
                and isinstance(result, (int, float)) and not isinstance(result, bool)
                and abs(result) > abs(threshold)
            ):
                issues.append("答案充分性.物理闭合与精度预算 结果超过阈值")
            digits = precision.get("报告有效位")
            if not isinstance(digits, int) or isinstance(digits, bool) or digits < 1:
                issues.append("答案充分性.物理闭合与精度预算.报告有效位 必须为正整数")
    if value.get("状态") != "PASS":
        issues.append("答案充分性.状态 must be PASS")
    return issues


def audit_metrics_schema(metrics: dict, require_precision_budget: bool = False) -> tuple[list[str], list[str]]:
    hard: list[str] = []
    warnings: list[str] = []
    results = metrics.get("关键结果")
    if not isinstance(results, list) or not results:
        hard.append("metrics.json lacks nonempty 关键结果 list")
    else:
        required_result = {"claim_id", "名称", "数值", "单位", "来源"}
        for index, item in enumerate(results, start=1):
            if not isinstance(item, dict):
                hard.append(f"关键结果 {index} is not an object")
                continue
            missing = sorted(required_result - set(item))
            if missing:
                hard.append(f"关键结果 {index} 缺字段: {', '.join(missing)}")

    validations = metrics.get("验证")
    if not isinstance(validations, list) or len(validations) < 2:
        hard.append("metrics.json requires at least two validation records")
    else:
        required_validation = {"类型", "指标", "数值", "阈值", "通过", "来源"}
        validation_types = set()
        for index, item in enumerate(validations, start=1):
            if not isinstance(item, dict):
                hard.append(f"验证 {index} is not an object")
                continue
            missing = sorted(required_validation - set(item))
            if missing:
                hard.append(f"验证 {index} 缺字段: {', '.join(missing)}")
            validation_types.add(str(item.get("类型", "")))
            if item.get("通过") is not True:
                hard.append(f"验证 {index} 未通过")
        if len(validation_types - {""}) < 2:
            hard.append("two validation records must use different mechanisms")

    if metrics.get("稳健性") in (None, "", [], {}):
        hard.append("metrics.json lacks robustness/sensitivity evidence")
    hard.extend(audit_model_selection(metrics.get("模型选择")))
    hard.extend(audit_answer_sufficiency(metrics.get("答案充分性")))
    sufficiency = metrics.get("答案充分性")
    if require_precision_budget and (
        not isinstance(sufficiency, dict) or "物理闭合与精度预算" not in sufficiency
    ):
        hard.append("机理/几何问题缺少答案充分性.物理闭合与精度预算")
    answer = metrics.get("最终回答")
    if not isinstance(answer, str) or not answer.strip():
        hard.append("metrics.json lacks nonempty 最终回答")
    if not validation_present(metrics):
        warnings.append("no validation-related top-level token detected")
    return hard, warnings


def audit_problem(problem_dir: Path, require_precision_budget: bool = False) -> dict:
    fig_dir = problem_dir / "图片"
    out_dir = problem_dir / "结果"
    figures = (
        [path for path in fig_dir.iterdir() if path.suffix.lower() in IMAGE_EXTS]
        if fig_dir.exists()
        else []
    )
    tables = (
        [
            path
            for path in out_dir.iterdir()
            if path.suffix.lower() in TABLE_EXTS and path.name != "metrics.json"
        ]
        if out_dir.exists()
        else []
    )
    scripts = sorted(problem_dir.glob("*.py"))
    metrics, metrics_status = load_json(out_dir / "metrics.json")

    hard: list[str] = []
    warnings: list[str] = []
    if not scripts:
        hard.append("reproducible main script missing")
    if not tables:
        hard.append("result table missing")
    if not figures:
        warnings.append("no result figure; confirm that the question genuinely needs none")
    if metrics is None:
        hard.append(f"metrics.json {metrics_status}")
    else:
        if not has_numeric_result(metrics):
            hard.append("metrics.json has no numerical result")
        schema_hard, schema_warnings = audit_metrics_schema(metrics, require_precision_budget)
        hard.extend(schema_hard)
        warnings.extend(schema_warnings)

    return {
        "problem": problem_dir.name,
        "scripts": len(scripts),
        "figures": len(figures),
        "tables": len(tables),
        "metrics": metrics_status,
        "pass": not hard,
        "hard_errors": hard,
        "warnings": warnings,
    }


def resolve_declared_path(root: Path, matrix_dir: Path, raw: str) -> bool:
    value = raw.split("#", 1)[0].strip()
    if not value or value.lower() in {"n/a", "none", "无"}:
        return False
    declared = Path(value)
    candidates = [declared] if declared.is_absolute() else [root / declared, matrix_dir / declared]
    for candidate in candidates:
        resolved = candidate.resolve()
        if (resolved == root or root in resolved.parents) and resolved.exists():
            return True
    return False


def audit_matrix(path: Path, root: Path) -> tuple[bool, list[str], int]:
    if not path.exists():
        return False, ["证据矩阵.csv missing"], 0
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            rows = list(csv.DictReader(handle))
    except Exception as exc:  # noqa: BLE001
        return False, [f"证据矩阵.csv unreadable: {exc}"], 0
    required = {
        "claim_id",
        "问题",
        "题目要求",
        "结论",
        "指标路径",
        "图表路径",
        "生成脚本",
        "验证方式",
        "模型选择依据",
        "充分性证据",
    }
    fields = set(rows[0].keys()) if rows else set()
    issues = []
    missing = sorted(required - fields)
    if missing:
        issues.append(f"证据矩阵缺列: {', '.join(missing)}")
    if not rows:
        issues.append("证据矩阵无记录")
    claim_ids = set()
    for index, row in enumerate(rows, start=2):
        for field in required:
            if not str(row.get(field, "")).strip():
                issues.append(f"证据矩阵第 {index} 行缺 {field}")
                break
        claim_id = str(row.get("claim_id", "")).strip()
        if claim_id in claim_ids:
            issues.append(f"证据矩阵 claim_id 重复: {claim_id}")
        claim_ids.add(claim_id)
        for field in ("指标路径", "图表路径", "生成脚本"):
            raw = str(row.get(field, "")).strip()
            if raw and not resolve_declared_path(root, path.parent, raw):
                issues.append(f"证据矩阵第 {index} 行 {field} 不存在: {raw}")
    return not issues, issues, len(rows)


def audit_contract(path: Path) -> tuple[dict | None, list[str]]:
    contract, status = load_json(path)
    if contract is None:
        return None, [f"任务契约.json {status}"]
    issues = []
    missing = sorted(CONTRACT_FIELDS - set(contract))
    if missing:
        issues.append(f"任务契约缺字段: {', '.join(missing)}")
    if contract.get("参考答案在求解期可见") is not False:
        issues.append("任务契约.参考答案在求解期可见 must be false")
    questions = contract.get("问题")
    if not isinstance(questions, list) or not questions:
        issues.append("任务契约问题列表为空")
        return contract, issues
    ids = []
    required = {
        "编号",
        "原型",
        "必答内容",
        "依赖",
        "答案证据要求",
        "主验证",
        "补充验证",
        "fallback",
        "失败条件",
    }
    for index, question in enumerate(questions, start=1):
        if not isinstance(question, dict):
            issues.append(f"任务契约问题 {index} 不是对象")
            continue
        absent = sorted(required - set(question))
        if absent:
            issues.append(f"任务契约问题 {index} 缺字段: {', '.join(absent)}")
        ids.append(question.get("编号"))
        for field in ("原型", "必答内容", "主验证", "补充验证", "fallback", "失败条件"):
            if not question.get(field):
                issues.append(f"任务契约问题 {index} 的 {field} 为空")
        answer_evidence = question.get("答案证据要求")
        if not isinstance(answer_evidence, list) or len(answer_evidence) < 2:
            issues.append(f"任务契约问题 {index} 的答案证据要求至少需要两项")
        routes = question.get("候选路线")
        unique_reason = str(question.get("唯一模型理由", "")).strip()
        if isinstance(routes, list) and len(routes) >= 2:
            route_required = {"名称", "角色", "关键假设", "淘汰条件"}
            route_names = []
            for route_index, route in enumerate(routes, start=1):
                if not isinstance(route, dict):
                    issues.append(f"任务契约问题 {index} 候选路线 {route_index} 不是对象")
                    continue
                absent_route = sorted(route_required - set(route))
                if absent_route:
                    issues.append(
                        f"任务契约问题 {index} 候选路线 {route_index} 缺字段: "
                        f"{', '.join(absent_route)}"
                    )
                route_names.append(str(route.get("名称", "")).strip())
                if not route.get("关键假设") or not route.get("淘汰条件"):
                    issues.append(f"任务契约问题 {index} 候选路线 {route_index} 缺假设或淘汰条件")
            if any(not name for name in route_names) or len(set(route_names)) != len(route_names):
                issues.append(f"任务契约问题 {index} 候选路线名称为空或重复")
        elif not unique_reason:
            issues.append(f"任务契约问题 {index} 需要至少两条候选路线或唯一模型理由")
        if not isinstance(question.get("依赖"), list):
            issues.append(f"任务契约问题 {index} 的依赖不是列表")
        elif any(not isinstance(dep, int) or dep >= index or dep < 1 for dep in question["依赖"]):
            issues.append(f"任务契约问题 {index} 存在非法前置依赖")
        failures = question.get("失败条件", [])
        if not isinstance(failures, list) or len(failures) < 2:
            issues.append(f"任务契约问题 {index} 至少需要两条失败条件")
        if question.get("主验证") == question.get("补充验证"):
            issues.append(f"任务契约问题 {index} 的两项验证机制相同")
    expected = list(range(1, len(questions) + 1))
    if ids != expected:
        issues.append(f"任务契约问题编号应为 {expected}，实际为 {ids}")
    return contract, issues


def imported_third_party(problem_dirs: list[Path]) -> set[str]:
    imported: set[str] = set()
    for problem_dir in problem_dirs:
        for script in problem_dir.glob("*.py"):
            try:
                tree = ast.parse(script.read_text(encoding="utf-8"))
            except Exception:  # noqa: BLE001
                continue
            for node in ast.walk(tree):
                if isinstance(node, ast.Import):
                    imported.update(alias.name.split(".", 1)[0] for alias in node.names)
                elif isinstance(node, ast.ImportFrom) and node.module:
                    imported.add(node.module.split(".", 1)[0])
    return set(THIRD_PARTY_IMPORTS) & imported


def audit_environment(
    path: Path, expected_questions: set[str], problem_dirs: list[Path]
) -> tuple[dict, list[str]]:
    environment, status = load_json(path)
    if environment is None:
        return {"status": status, "python_version": None, "packages": 0, "run_commands": 0}, [
            f"运行环境.json {status}"
        ]

    issues: list[str] = []
    python_version = environment.get("python_version")
    if not isinstance(python_version, str) or not python_version.strip():
        issues.append("运行环境.json 缺 python_version")
    packages = environment.get("packages")
    if not isinstance(packages, dict) or not packages:
        issues.append("运行环境.json 缺非空 packages")
        packages = {}
    elif any(not str(name).strip() or not str(version).strip() for name, version in packages.items()):
        issues.append("运行环境.json packages 含空名称或版本")
    normalized_packages = {str(name).strip().lower().replace("_", "-") for name in packages}
    for module in sorted(imported_third_party(problem_dirs)):
        aliases = {name.lower().replace("_", "-") for name in THIRD_PARTY_IMPORTS[module]}
        if normalized_packages.isdisjoint(aliases):
            issues.append(f"运行环境.json packages 漏记导入依赖: {module}")
    commands = environment.get("run_commands")
    if not isinstance(commands, dict):
        issues.append("运行环境.json run_commands 必须为对象")
        commands = {}
    else:
        missing = sorted(expected_questions - set(commands))
        if missing:
            issues.append(f"运行环境.json 缺逐问命令: {', '.join(missing)}")
        for question, command in commands.items():
            if not isinstance(command, str) or not command.strip():
                issues.append(f"运行环境.json 的 {question} 命令为空")
    if environment.get("verified") is not True:
        issues.append("运行环境.json verified 不是 true")
    return {
        "status": "ok" if not issues else "invalid",
        "python_version": python_version,
        "packages": len(packages),
        "run_commands": len(commands),
    }, issues


def workbook_output_cells(path: Path) -> tuple[int | None, str | None]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError:
        return None, "openpyxl unavailable"
    try:
        workbook = load_workbook(path, read_only=True, data_only=False)
        count = 0
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    value = cell.value
                    if isinstance(value, bool):
                        continue
                    if isinstance(value, (int, float)):
                        count += 1
                    elif isinstance(value, str) and value.startswith("="):
                        count += 1
        workbook.close()
        return count, None
    except Exception as exc:  # noqa: BLE001
        return None, str(exc)


def audit_official_outputs(root: Path, contract: dict | None) -> tuple[list[dict], list[str]]:
    results = []
    issues = []
    if not contract:
        return results, issues
    for index, item in enumerate(contract.get("官方输出", []), start=1):
        if not isinstance(item, dict) or not item.get("路径"):
            continue
        raw = Path(str(item["路径"]))
        candidates = [raw] if raw.is_absolute() else [root / raw, root / "数据" / raw.name]
        path = next((candidate.resolve() for candidate in candidates if candidate.exists()), None)
        record = {"declared": str(raw), "resolved": str(path) if path else None, "status": "ok"}
        if path is None:
            record["status"] = "missing"
            issues.append(f"official output missing: {raw}")
        elif path != root and root not in path.parents:
            record["status"] = "outside-project"
            issues.append(f"official output resolves outside project: {raw}")
        elif path.suffix.lower() == ".xlsx":
            cells, error = workbook_output_cells(path)
            record["numeric_or_formula_cells"] = cells
            if error:
                record["status"] = "unreadable"
                issues.append(f"official output unreadable {raw}: {error}")
            elif cells == 0:
                record["status"] = "empty"
                issues.append(f"official output has no numeric/formula cells: {raw}")
        results.append(record)
    return results, issues


def write_report(root: Path, result: dict) -> None:
    solve_dir = root / "求解"
    solve_dir.mkdir(exist_ok=True)
    (solve_dir / "证据审计.json").write_text(
        json.dumps(result, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    lines = ["# 求解证据审计", "", f"结论：{'PASS' if result['pass'] else 'FAIL'}", ""]
    for item in result["problems"]:
        lines.extend(
            [
                f"## {item['problem']}",
                f"- 主脚本：{item['scripts']}",
                f"- 图片：{item['figures']}",
                f"- 结果表：{item['tables']}",
                f"- metrics.json：{item['metrics']}",
            ]
        )
        lines.extend(f"- 硬错误：{issue}" for issue in item["hard_errors"])
        lines.extend(f"- 警告：{warning}" for warning in item["warnings"])
        lines.append("")
    lines.append("## 全局检查")
    lines.append(f"- 任务契约问题数：{result['contract_questions']}")
    lines.append(f"- 官方输出检查数：{len(result['official_outputs'])}")
    for output in result["official_outputs"]:
        lines.append(f"  - {output['declared']}: {output['status']}")
    lines.append(f"- 证据矩阵记录：{result['matrix_rows']}")
    lines.append(
        f"- 运行环境：{result['environment']['status']}，Python {result['environment']['python_version']}，"
        f"包 {result['environment']['packages']}，逐问命令 {result['environment']['run_commands']}"
    )
    lines.extend(f"- {issue}" for issue in result["global_issues"])
    if not result["global_issues"]:
        lines.append("- 无")
    (solve_dir / "证据审计.md").write_text("\n".join(lines), encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--root", default=".", help="CUMCM project root")
    parser.add_argument("--no-write", action="store_true")
    args = parser.parse_args()

    root = Path(args.root).resolve()
    solve_dir = root / "求解"
    problem_dirs = sorted(path for path in solve_dir.glob("问题*") if path.is_dir())

    global_issues: list[str] = []
    contract, contract_issues = audit_contract(solve_dir / "任务契约.json")
    global_issues.extend(contract_issues)
    official_outputs, output_issues = audit_official_outputs(root, contract)
    global_issues.extend(output_issues)
    if not (solve_dir / "求解计划.md").exists():
        global_issues.append("求解计划.md missing")
    if not (solve_dir / "图表清单.md").exists():
        global_issues.append("图表清单.md missing")
    if not problem_dirs:
        global_issues.append("no 求解/问题X directories")
    contract_questions = len(contract.get("问题", [])) if contract else 0
    expected_dirs = {f"问题{index}" for index in range(1, contract_questions + 1)}
    actual_dirs = {path.name for path in problem_dirs}
    if expected_dirs and actual_dirs != expected_dirs:
        global_issues.append(
            f"问题目录与任务契约不一致: expected={sorted(expected_dirs)}, actual={sorted(actual_dirs)}"
        )
    environment, environment_issues = audit_environment(
        solve_dir / "运行环境.json", expected_dirs, problem_dirs
    )
    global_issues.extend(environment_issues)
    matrix_ok, matrix_issues, matrix_rows = audit_matrix(solve_dir / "证据矩阵.csv", root)
    if not matrix_ok:
        global_issues.extend(matrix_issues)

    prototype_by_dir: dict[str, set[str]] = {}
    if contract:
        for question in contract.get("问题", []):
            if not isinstance(question, dict) or not isinstance(question.get("编号"), int):
                continue
            prototype_by_dir[f"问题{question['编号']}"] = {
                str(item) for item in question.get("原型", [])
            }
    precision_prototypes = {"机理", "几何", "动力学", "控制"}
    problems = [
        audit_problem(
            path,
            bool(prototype_by_dir.get(path.name, set()) & precision_prototypes),
        )
        for path in problem_dirs
    ]
    passed = not global_issues and bool(problems) and all(item["pass"] for item in problems)
    result = {
        "pass": passed,
        "global_issues": global_issues,
        "contract_questions": contract_questions,
        "official_outputs": official_outputs,
        "environment": environment,
        "matrix_rows": matrix_rows,
        "problems": problems,
    }
    if args.no_write:
        print(json.dumps(result, ensure_ascii=False))
    else:
        write_report(root, result)

    print("PASS" if passed else "FAIL")
    if not args.no_write:
        print(f"Wrote {solve_dir / '证据审计.md'}")
    return 0 if passed else 1


if __name__ == "__main__":
    raise SystemExit(main())
